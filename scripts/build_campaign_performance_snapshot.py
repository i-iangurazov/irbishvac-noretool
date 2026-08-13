from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


COMPANY_REVENUE_GOAL = 2_000_000
MARKETING_BUDGET_RATE = 0.07
CAPACITY_ASSUMPTIONS = (
    {"team": "HVAC Service", "people": 5, "opportunitiesPerDay": 3},
    {"team": "Maintenance", "people": 2, "opportunitiesPerDay": 3},
    {"team": "Commercial", "people": 1, "opportunitiesPerDay": 3},
    {"team": "Plumbing Service", "people": 3, "opportunitiesPerDay": 3},
    {"team": "Comfort Advisors", "people": 3, "opportunitiesPerDay": 4},
)
PLANNING_DAYS = 25


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build a normalized campaign Plan vs Actual snapshot."
    )
    parser.add_argument("--current-call-center", required=True)
    parser.add_argument("--prior-call-center", required=True)
    parser.add_argument("--campaign-summary", required=True)
    parser.add_argument("--prior-campaign-summary", required=True)
    parser.add_argument("--sold-estimates", required=True)
    parser.add_argument("--from-date", required=True)
    parser.add_argument("--to-date", required=True)
    parser.add_argument("--period-label")
    parser.add_argument("--plan-status", default="DRAFT")
    parser.add_argument("--plan-reference-label")
    parser.add_argument("--planning-note")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    start = date.fromisoformat(args.from_date)
    cutoff = date.fromisoformat(args.to_date)
    period_label = args.period_label or cutoff.strftime("%B %Y MTD")
    plan_reference_label = args.plan_reference_label or "prior-period"
    planning_note = args.planning_note or (
        f"Draft allocation by each channel's share of {plan_reference_label} qualified leads."
    )
    current = _read_call_center(Path(args.current_call_center), cutoff)
    prior = _read_call_center(Path(args.prior_call_center), None)
    campaign_actual = _read_campaign_summary(Path(args.campaign_summary))
    prior_campaign = _read_campaign_summary(Path(args.prior_campaign_summary))
    sales_actual = _read_sold_estimates(Path(args.sold_estimates))

    lead_goal = sum(
        row["people"] * row["opportunitiesPerDay"] * PLANNING_DAYS
        for row in CAPACITY_ASSUMPTIONS
    )
    elapsed_ratio = cutoff.day / _days_in_month(cutoff)
    prior_leads_by_channel = {
        name: row["qualifiedLeads"] for name, row in prior["channels"].items()
    }
    lead_goals = _allocate_integer_goal(lead_goal, prior_leads_by_channel)
    prior_revenue_by_channel = {
        name: row["completedRevenue"] for name, row in prior_campaign.items()
    }
    budget_goals = _allocate_float_goal(
        COMPANY_REVENUE_GOAL * MARKETING_BUDGET_RATE,
        prior_revenue_by_channel,
    )
    revenue_goals = _allocate_float_goal(
        COMPANY_REVENUE_GOAL,
        prior_revenue_by_channel,
    )

    channel_names = set(current["channels"])
    channel_names.update(campaign_actual)
    channel_names.update(sales_actual)
    channel_names.update(lead_goals)
    channel_names.update(budget_goals)
    channel_names.discard("Other")

    rows = []
    for channel in channel_names:
        calls = current["channels"].get(channel, _empty_call_center_row())
        campaign = campaign_actual.get(channel, _empty_campaign_row())
        sales = sales_actual.get(channel, {"soldJobs": 0, "soldAmount": 0.0})
        monthly_lead_goal = lead_goals.get(channel, 0)
        prior_calls = prior["channels"].get(channel, _empty_call_center_row())
        prior_booking_rate = _ratio(
            prior_calls["bookedJobs"], prior_calls["qualifiedLeads"]
        )
        lead_attainment = (
            calls["qualifiedLeads"] / monthly_lead_goal if monthly_lead_goal else None
        )
        pace = lead_attainment / elapsed_ratio if lead_attainment is not None else None
        rows.append(
            {
                "channel": channel,
                "plan": {
                    "qualifiedLeads": monthly_lead_goal,
                    "bookedJobs": (
                        round(monthly_lead_goal * prior_booking_rate)
                        if prior_booking_rate is not None
                        else None
                    ),
                    "spend": budget_goals.get(channel),
                    "soldAmount": None,
                    "completedRevenue": revenue_goals.get(channel),
                },
                "actual": {
                    "calls": calls["calls"],
                    "forms": calls["forms"],
                    "qualifiedLeads": calls["qualifiedLeads"],
                    "bookedJobs": calls["bookedJobs"],
                    "bookingRate": _ratio(calls["bookedJobs"], calls["qualifiedLeads"]),
                    "spend": campaign["spend"],
                    "costPerLead": _ratio(campaign["spend"], calls["qualifiedLeads"]),
                    "soldJobs": sales["soldJobs"],
                    "soldAmount": sales["soldAmount"],
                    "completedRevenue": campaign["completedRevenue"],
                    "roi": _roi(campaign["completedRevenue"], campaign["spend"]),
                },
                "leadAttainment": lead_attainment,
                "pace": pace,
                "status": _status(pace, campaign["spend"], sales["soldJobs"]),
            }
        )

    rows.sort(
        key=lambda row: (
            -row["actual"]["qualifiedLeads"],
            -row["actual"]["completedRevenue"],
            row["channel"],
        )
    )

    totals = _totals(rows)
    alerts = _alerts(rows)
    payload = {
        "schemaVersion": 1,
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "period": {
            "id": start.strftime("%Y-%m"),
            "label": period_label,
            "from": start.isoformat(),
            "to": cutoff.isoformat(),
            "elapsedCalendarDays": cutoff.day,
            "calendarDaysInMonth": _days_in_month(cutoff),
        },
        "plan": {
            "status": args.plan_status,
            "companyRevenueGoal": COMPANY_REVENUE_GOAL,
            "marketingBudgetRate": MARKETING_BUDGET_RATE,
            "marketingBudgetGoal": COMPANY_REVENUE_GOAL * MARKETING_BUDGET_RATE,
            "qualifiedLeadGoal": lead_goal,
            "planningDays": PLANNING_DAYS,
            "capacityAssumptions": CAPACITY_ASSUMPTIONS,
            "channelLeadGoalMethod": (
                planning_note
            ),
            "channelBudgetGoalStatus": (
                f"Model allocation by {plan_reference_label} completed revenue share"
            ),
        },
        "actual": totals,
        "pace": {
            "expectedToDateRatio": elapsed_ratio,
            "qualifiedLeadPace": _ratio(
                totals["qualifiedLeads"], lead_goal * elapsed_ratio
            ),
            "spendPace": _ratio(
                totals["spend"],
                COMPANY_REVENUE_GOAL * MARKETING_BUDGET_RATE * elapsed_ratio,
            ),
        },
        "alerts": alerts,
        "rows": rows,
        "sources": [
            {
                "name": f"{start.strftime('%B %Y')} Call Center Report",
                "role": "Calls, forms, qualified leads, booked jobs",
                "cutoffApplied": cutoff.isoformat(),
            },
            {
                "name": "ServiceTitan Campaign Summary Report",
                "reportId": "898",
                "role": "Spend and completed revenue",
            },
            {
                "name": "ServiceTitan Sold Estimates",
                "reportId": "7148368",
                "role": "Sold jobs and sold amount by parent job campaign",
            },
            {
                "name": "ServiceTitan Revenue By Campaign",
                "reportId": "101394656",
                "role": "Campaign revenue verification",
            },
        ],
        "dataNotes": [
            "Call Center rows dated after the MTD cutoff are excluded.",
            "Calls and forms are combined as leads on the executive dashboard.",
            "Granular ServiceTitan campaign names are normalized to the call-center channels.",
            planning_note,
        ],
    }

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"wrote {output}")
    print(
        f"actual leads={totals['qualifiedLeads']} booked={totals['bookedJobs']} "
        f"sold={totals['soldJobs']} sold_amount={totals['soldAmount']:.2f} "
        f"revenue={totals['completedRevenue']:.2f} spend={totals['spend']:.2f}"
    )
    return 0


def _read_call_center(path: Path, cutoff: date | None) -> dict[str, Any]:
    sheet = openpyxl.load_workbook(path, data_only=True)["Master Sheet"]
    channels: dict[str, dict[str, int]] = defaultdict(_empty_call_center_row)
    for row in range(2, sheet.max_row + 1):
        received = sheet.cell(row, 4).value
        if not isinstance(received, datetime):
            continue
        if cutoff and received.date() > cutoff:
            continue
        channel = _normalize_channel(str(sheet.cell(row, 14).value or "Miscellaneous"))
        medium = str(sheet.cell(row, 7).value or "")
        quality = str(sheet.cell(row, 8).value or "")
        stage = str(sheet.cell(row, 9).value or "")
        current = channels[channel]
        current["requests"] += 1
        current["calls"] += int(medium == "Call")
        current["forms"] += int(medium == "Text")
        current["qualifiedLeads"] += int(quality in {"Good", "Mid"})
        current["bookedJobs"] += int(stage == "Booked")
    return {"channels": dict(channels)}


def _read_campaign_summary(path: Path) -> dict[str, dict[str, float]]:
    channels: dict[str, dict[str, float]] = defaultdict(_empty_campaign_row)
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            channel = _normalize_channel(row.get("Name") or "")
            if channel == "Other":
                continue
            current = channels[channel]
            current["reportLeads"] += _number(row.get("Leads"))
            current["reportBookedJobs"] += _number(row.get("TotalJobsBooked"))
            current["spend"] += _number(row.get("Cost"))
            current["completedRevenue"] += _number(row.get("CompletedRevenue"))
    return dict(channels)


def _read_sold_estimates(path: Path) -> dict[str, dict[str, float]]:
    channels: dict[str, dict[str, float]] = defaultdict(
        lambda: {"soldJobs": 0, "soldAmount": 0.0}
    )
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            channel = _normalize_channel(row.get("ParentJobCampaign") or "")
            if channel == "Other":
                continue
            channels[channel]["soldJobs"] += 1
            channels[channel]["soldAmount"] += _number(row.get("Total"))
    return dict(channels)


def _normalize_channel(value: str) -> str:
    normalized = re.sub(r"\s+", " ", value).strip().lower()
    if not normalized:
        return "Miscellaneous"
    rules = (
        ("Yelp", ("yelp",)),
        ("Google LSA", ("google local services", "google lsa", "lsa")),
        ("Website", ("direct web traffic", "google organic", "website")),
        ("GBP San Jose", ("gbp san jose", "google business")),
        ("Google Ads", ("google ads", "maxconv", "pmax", "irbis |")),
        ("Facebook", ("facebook", "paid social", "social")),
        ("Hatch Campaigns", ("hatch",)),
        ("Workfuel", ("workfuel", "work fuel")),
        ("Carrier", ("carrier",)),
        ("669-COOLING", ("669-cooling", "669 cooling")),
        ("Mail Shark", ("mail shark", "direct mail")),
        ("Scheduling Pro", ("scheduling pro",)),
        ("Home Care Plan", ("home care plan",)),
        ("Existing Customers", ("existing customer",)),
        ("Now Operator", ("now operator",)),
        ("Miscellaneous", ("miscellaneous", "recall", "warranty")),
    )
    for label, markers in rules:
        if any(marker in normalized for marker in markers):
            return label
    return "Other"


def _allocate_integer_goal(total: int, weights: dict[str, int]) -> dict[str, int]:
    relevant = {name: value for name, value in weights.items() if value > 0 and name != "Other"}
    weight_total = sum(relevant.values())
    if not weight_total:
        return {}
    raw = {name: total * weight / weight_total for name, weight in relevant.items()}
    allocated = {name: math.floor(value) for name, value in raw.items()}
    remainder = total - sum(allocated.values())
    order = sorted(raw, key=lambda name: raw[name] - allocated[name], reverse=True)
    for name in order[:remainder]:
        allocated[name] += 1
    return allocated


def _allocate_float_goal(total: float, weights: dict[str, float]) -> dict[str, float]:
    relevant = {
        name: value for name, value in weights.items() if value > 0 and name != "Other"
    }
    weight_total = sum(relevant.values())
    if not weight_total:
        return {}
    return {
        name: round(total * weight / weight_total, 2)
        for name, weight in relevant.items()
    }


def _totals(rows: list[dict[str, Any]]) -> dict[str, float]:
    metrics = (
        "calls",
        "forms",
        "qualifiedLeads",
        "bookedJobs",
        "spend",
        "soldJobs",
        "soldAmount",
        "completedRevenue",
    )
    totals = {
        metric: sum(float(row["actual"][metric]) for row in rows) for metric in metrics
    }
    totals["bookingRate"] = _ratio(totals["bookedJobs"], totals["qualifiedLeads"])
    totals["costPerLead"] = _ratio(totals["spend"], totals["qualifiedLeads"])
    totals["roi"] = _roi(totals["completedRevenue"], totals["spend"])
    return totals


def _alerts(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    alerts = []
    for row in rows:
        actual = row["actual"]
        if actual["spend"] >= 500 and actual["soldJobs"] == 0:
            alerts.append(
                {
                    "severity": "critical",
                    "channel": row["channel"],
                    "message": f"${actual['spend']:,.0f} spent with no sold estimates MTD.",
                }
            )
        if actual["qualifiedLeads"] >= 5 and actual["bookingRate"] is not None and actual["bookingRate"] < 0.5:
            alerts.append(
                {
                    "severity": "warning",
                    "channel": row["channel"],
                    "message": f"Booking rate is {actual['bookingRate']:.0%} on {actual['qualifiedLeads']:.0f} qualified leads.",
                }
            )
    return sorted(alerts, key=lambda row: row["severity"] == "warning")[:4]


def _status(pace: float | None, spend: float, sold_jobs: int) -> str:
    if spend >= 500 and sold_jobs == 0:
        return "risk"
    if pace is None:
        return "unplanned"
    if pace >= 1:
        return "on-track"
    if pace >= 0.8:
        return "watch"
    return "off-track"


def _empty_call_center_row() -> dict[str, int]:
    return {"requests": 0, "calls": 0, "forms": 0, "qualifiedLeads": 0, "bookedJobs": 0}


def _empty_campaign_row() -> dict[str, float]:
    return {"reportLeads": 0.0, "reportBookedJobs": 0.0, "spend": 0.0, "completedRevenue": 0.0}


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _ratio(numerator: float, denominator: float) -> float | None:
    return numerator / denominator if denominator else None


def _roi(revenue: float, spend: float) -> float | None:
    return (revenue - spend) / spend if spend else None


def _days_in_month(value: date) -> int:
    if value.month == 12:
        next_month = date(value.year + 1, 1, 1)
    else:
        next_month = date(value.year, value.month + 1, 1)
    return (next_month - date(value.year, value.month, 1)).days


if __name__ == "__main__":
    raise SystemExit(main())
